import pandas as pd

# Load CSV files
questionnaire = pd.read_csv("questionnaire.csv", low_memory=False)
utilisation_du_sol = pd.read_csv("utilisation_du_sol.csv")
materiel_agricole = pd.read_csv("materiel_agricole.csv")
post_superficie_exploitation = pd.read_csv("post_superficie_exploitation.csv")
status_juridique = pd.read_csv("status_juridique.csv")

# ✅ Remove duplicate id_questionnaire values (Keep first entry)
utilisation_du_sol = utilisation_du_sol.drop_duplicates(subset=["id_questionnaire"], keep="first")
materiel_agricole = materiel_agricole.drop_duplicates(subset=["id_questionnaire"], keep="first")
status_juridique = status_juridique.drop_duplicates(subset=["id_questionnaire"], keep="first")

# ✅ Aggregate numeric values (e.g., mean, sum)
post_superficie_exploitation = post_superficie_exploitation.groupby("id_questionnaire").agg("sum").reset_index()

# ✅ Merge data on 'id_questionnaire' using a left join
merged_data = questionnaire.copy()
merged_data = pd.merge(merged_data, utilisation_du_sol, on="id_questionnaire", how="left")
merged_data = pd.merge(merged_data, materiel_agricole, on="id_questionnaire", how="left")
merged_data = pd.merge(merged_data, post_superficie_exploitation, on="id_questionnaire", how="left")
merged_data = pd.merge(merged_data, status_juridique, on="id_questionnaire", how="left")

# ✅ Pivot transformation
grouped_data = merged_data.groupby("id_questionnaire").agg(lambda x: list(x))

# ✅ Reformat data: Convert lists to dynamic columns
transformed_rows = []
for idx, row in grouped_data.iterrows():
    new_row = {
        "id_questionnaire": idx,
        "exploitant_cle_unique": row["exploitant_cle_unique"][0] if "exploitant_cle_unique" in row else "",
        "origine_des_terres": row["origine_des_terres"][0] if "origine_des_terres" in row else "",
        "status_juridique": row["status_juridique"][0] if "status_juridique" in row else "",
        "superfecie_sj": row["superfecie_sj"][0] if "superfecie_sj" in row else "",
        "superfecie_sj_are": row["superfecie_sj_are"][0] if "superfecie_sj_are" in row else "",
    }

    # ✅ Handling cultures dynamically
    for i, (culture, superficie, superficie_are) in enumerate(zip(
        row.get("code_culture", []), 
        row.get("superficie_hec", []), 
        row.get("superficie_are", [])
    ), start=1):
        new_row[f"code_culture{i}"] = culture
        new_row[f"superficie_hec{i}"] = superficie
        new_row[f"superficie_are{i}"] = superficie_are

    # ✅ Handling materials dynamically
    for i, (materiel, nombre, mode_mobilisation, mode_exploitation) in enumerate(zip(
        row.get("code_materiel", []), 
        row.get("code_materiel_nombre", []), 
        row.get("ee_mode_mobilisation_materiel", []), 
        row.get("ee_mode_exploitation_materiel", [])
    ), start=1):
        new_row[f"code_materiel{i}"] = materiel
        new_row[f"code_materiel_nombre{i}"] = nombre
        new_row[f"ee_mode_mobilisation_materiel{i}"] = mode_mobilisation
        new_row[f"ee_mode_exploitation_materiel{i}"] = mode_exploitation

    # ✅ Handling superficies dynamically
    for i, (sau, sat, st) in enumerate(zip(
        row.get("superficie_agricole_utile_sau_1", []),
        row.get("superficie_agricole_totale_sat_1", []),
        row.get("surface_totale_st_1", [])
    ), start=1):
        new_row[f"superficie_agricole_utile_sau_{i}"] = sau
        new_row[f"superficie_agricole_totale_sat_{i}"] = sat
        new_row[f"surface_totale_st_{i}"] = st

    transformed_rows.append(new_row)

# ✅ Convert back to DataFrame
final_df = pd.DataFrame(transformed_rows)

# ✅ Export to Excel
output_file = "questionnaire_transformed.xlsx"
final_df.to_excel(output_file, index=False)

print(f"✅ Transformation successful! File saved as {output_file}")

--------------------------------------------------------------------------------------------lastly made----------------------------------------------
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

# Load data
df = pd.read_csv("recenseur.csv", low_memory=False)

# Rename columns with French labels
df.rename(columns={
    "id_recensseur": "ID Recenseur",
    "id_user": "ID Utilisateur",
    "nom_recensseur": "Nom Recenseur",
    "prenom_recenseur": "Prénom Recenseur",
    "commune": "Commune (Code)",
    "email": "Email",
    "controleur": "Contrôleur",
    "phone": "Téléphone",
    "num_zone_district": "Numéro Zone District",
    "num_exploitation": "Numéro Exploitation",
    "nom_zone_district": "Nom Zone District",
    "creation_date": "Date de Création",
    "id": "ID",
    "commune_code": "Code Commune",
    "commune_name": "Nom de la Commune",
    "commune_name_ascii": "Nom Commune (ASCII)",
    "daira_name": "Nom Daira",
    "daira_name_ascii": "Nom Daira (ASCII)",
    "wilaya_code": "Code Wilaya",
    "wilaya_name": "Nom Wilaya",
    "wilaya_name_ascii": "Nom Wilaya (ASCII)",
    "qst_a_recense": "Questionnaires à Recenser",
    "qst_recense": "Questionnaires Renseignés"
}, inplace=True)

# Export to Excel
output_file = "questionnaire_transformed3.xlsx"
df.to_excel(output_file, index=False)

# Optional: Apply header style (bold + fill color)
wb = load_workbook(output_file)
ws = wb.active

# Style headers
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

for cell in ws[1]:
    cell.font = header_font
    cell.fill = header_fill

# Save styled file
wb.save(output_file)

print(f"✅ Export terminé avec succès : {output_file}")

