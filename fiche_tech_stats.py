import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

# Load data
questionnaire = pd.read_csv("questionnaire.csv", low_memory=False)
communes = pd.read_csv("communes.csv")

# Merge data on 'commune_code' using a LEFT JOIN
merged_data = pd.merge(
    questionnaire, communes[['commune_code', 'wilaya_name_ascii']], 
    on="commune_code", 
    how="left"
)

# Rename columns for clarity
merged_data = merged_data.rename(columns={
    "id_questionnaire": "ID Questionnaire",
    "exploitant_cle_unique": "Exploitant Clé Unique",
    "commune_code": "Commune Code",
    "wilaya_name_ascii": "Wilaya Name",
    "f_phone_exploitant": "Numéro de téléphone",
    "f_recensement_question_check_oui": "Saviez-vous qu'il a eu lieu le Recensement Général",
    "f_recensement_si_check_non": "Êtes-vous recensé ?",
    "f_date_passage": "Si oui, Quelle date ?",
    "f_reason": "Raison de la non-participation au recensement",
    "f_post_recensement": "Enquête post-censitaire"
})

# Ensure 'Commune Code' is formatted with leading zeros
merged_data["Commune Code"] = (
    merged_data["Commune Code"]
    .fillna(0)
    .astype(int)
    .astype(str)
    .str.zfill(4)
)

# Export data to Excel
output_file = "بطاقة التحقيق.xlsx"
merged_data.to_excel(output_file, index=False, sheet_name="Data")

# Load workbook
wb = load_workbook(output_file)

# Add a statistics sheet
if "Statistics" not in wb.sheetnames:
    wb.create_sheet("Statistics")
ws_stats = wb["Statistics"]

# Calculate statistics
total_questionnaires = len(merged_data)
questionnaires_per_wilaya = merged_data["Wilaya Name"].value_counts()
participation_rate = merged_data["Êtes-vous recensé ?"].value_counts()
non_participation_reasons = merged_data["Raison de la non-participation au recensement"].value_counts()

# Write statistics to sheet
ws_stats.append(["Statistics Summary"])
ws_stats.append(["Total Questionnaires Collected", total_questionnaires])
ws_stats.append([""])

ws_stats.append(["Questionnaires per Wilaya"])
for wilaya, count in questionnaires_per_wilaya.items():
    ws_stats.append([wilaya, count])
ws_stats.append([""])

ws_stats.append(["Participation Rate"])
for response, count in participation_rate.items():
    ws_stats.append([response, count])
ws_stats.append([""])

ws_stats.append(["Reasons for Non-Participation"])
for reason, count in non_participation_reasons.items():
    ws_stats.append([reason, count])

# Apply formatting
header_font = Font(bold=True)
for cell in ws_stats["A"]:
    cell.font = header_font

# Save updated file
wb.save(output_file)

print(f"✅ Statistics added! Updated file saved as {output_file}")


# f_phone_exploitant   Numéro de téléphone 
# f_recensement_question_check_oui Saviez-vous qu'il a eu lieu le Recensement Général 
# f_recensement_si_check_non  Êtes-vous recensé ?
# f_day_of_passage Si oui, Quelle jour ?
# f_reason   Raison de la non-participation au recensement 
# f_post_recensement   Enquête post-censitaire 


# SELECT 
#     q.id_questionnaire AS "ID Questionnaire", 
#     q.exploitant_cle_unique AS "Exploitant Clé Unique", 
#     q.commune_code AS "Commune Code", 
#     c.wilaya_name_ascii AS "Wilaya Name", 
#     q.f_phone_exploitant AS "Numéro de téléphone", 
#     q.f_recensement_question_check_oui AS "Saviez-vous qu'il a eu lieu le Recensement Général", 
#     q.f_recensement_si_check_non AS "Êtes-vous recensé ?", 
#     q.f_date_passage AS "Si oui, Quelle date ?", 
#     q.f_reason AS "Raison de la non-participation au recensement", 
#     q.f_post_recensement AS "Enquête post-censitaire"
# FROM questionnaire q
# LEFT JOIN communes c ON q.commune_code = c.commune_code;


# from questionnaire; 

# SELECT wilaya_name FROM `communes` WHERE `commune_code`=""
# SELECT wilaya_name_ascii, commune_code FROM `communes` WHERE 1;