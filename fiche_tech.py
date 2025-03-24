import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

# Load data
questionnaire = pd.read_csv("questionnaire.csv", low_memory=False)
communes = pd.read_csv("communes.csv")

# Merge data on 'commune_code' using a LEFT JOIN
merged_data = pd.merge(
    questionnaire, communes[['commune_code', 'wilaya_name_ascii']], 
    on="commune_code", 
    how="left"
)

# Rename columns for clarity
merged_data = merged_data.rename(columns={
    "id_questionnaire": "ID Questionnaire",
    "exploitant_cle_unique": "Exploitant Clé Unique",
    "commune_code": "Commune Code",
    "wilaya_name_ascii": "Wilaya Name",
    "f_phone_exploitant": "Numéro de téléphone",
    "f_recensement_question_check_oui": "Saviez-vous qu'il a eu lieu le Recensement Général",
    "f_recensement_si_check_non": "Êtes-vous recensé ?",
    "f_date_passage": "Si oui, Quelle date ?",
    "f_reason": "Raison de la non-participation au recensement",
    "f_post_recensement": "Enquête post-censitaire"
})

# Ensure 'Commune Code' is formatted with leading zeros
merged_data["Commune Code"] = (
    merged_data["Commune Code"]
    .fillna(0)  # Replace NaN with 0
    .astype(int)  # Convert to integer
    .astype(str)  # Convert to string
    .str.zfill(4)  # Ensure 4-digit format
)

# Export to Excel without formatting
output_file = "بطاقة التحقيق.xlsx"
merged_data.to_excel(output_file, index=False, sheet_name="Data")

# Load workbook and get active sheet
wb = load_workbook(output_file)
ws = wb.active

# Define styles
header_font = Font(bold=True)  # Bold font
header_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")  # Light green background

# Apply styles to header row
for col in ws.iter_cols(min_row=1, max_row=1, min_col=1, max_col=ws.max_column):
    for cell in col:
        cell.font = header_font
        cell.fill = header_fill

# Adjust row height for header
ws.row_dimensions[1].height = 25  # Set header row height

# Save the formatted file
wb.save(output_file)

print(f"✅ Merge successful! Formatted file saved as {output_file}")

# f_phone_exploitant   Numéro de téléphone 
# f_recensement_question_check_oui Saviez-vous qu'il a eu lieu le Recensement Général 
# f_recensement_si_check_non  Êtes-vous recensé ?
# f_day_of_passage Si oui, Quelle jour ?
# f_reason   Raison de la non-participation au recensement 
# f_post_recensement   Enquête post-censitaire 


# SELECT 
#     q.id_questionnaire AS "ID Questionnaire", 
#     q.exploitant_cle_unique AS "Exploitant Clé Unique", 
#     q.commune_code AS "Commune Code", 
#     c.wilaya_name_ascii AS "Wilaya Name", 
#     q.f_phone_exploitant AS "Numéro de téléphone", 
#     q.f_recensement_question_check_oui AS "Saviez-vous qu'il a eu lieu le Recensement Général", 
#     q.f_recensement_si_check_non AS "Êtes-vous recensé ?", 
#     q.f_date_passage AS "Si oui, Quelle date ?", 
#     q.f_reason AS "Raison de la non-participation au recensement", 
#     q.f_post_recensement AS "Enquête post-censitaire"
# FROM questionnaire q
# LEFT JOIN communes c ON q.commune_code = c.commune_code;


# from questionnaire; 

# SELECT wilaya_name FROM `communes` WHERE `commune_code`=""
# SELECT wilaya_name_ascii, commune_code FROM `communes` WHERE 1;